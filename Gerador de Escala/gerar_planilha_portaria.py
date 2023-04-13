from openpyxl import Workbook
from openpyxl.drawing.image import Image 
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Color, Font, Border, Side, PatternFill, fills, borders
from datetime import datetime, timedelta, date

#configurando bordas
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

#capturando dados de data 
data_atual = date.today()
year = data_atual.year

#criando arquivo 
wb = Workbook()

#selecionando a planilha atual
ws = wb.active

#mudando o titulo da folha atual
ws.title = "Portaria {}".format(2023)

#carregar imagem
img = Image("logo.png")
img.height = 75
img.width = 180

#adicionar imagem a celula A1 da planilha
ws.add_image(img, "A1")

#mesclar celulas para imagem e titulo
ws.merge_cells("A1:C4")
ws.merge_cells("D1:K4")
ws["D1"].alignment = Alignment(horizontal="center")

#meclar celulas cabeçalho e ajustar alinhamento
ws.merge_cells("A5:B5")
ws["A5"].alignment = Alignment(horizontal="center")
ws.merge_cells("C5:D5")
ws["C5"].alignment = Alignment(horizontal="center")
ws.merge_cells("E5:F5")
ws["E5"].alignment = Alignment(horizontal="center")
ws.merge_cells("G5:H5")
ws["G5"].alignment = Alignment(horizontal="center")
ws.merge_cells("I5:K5")
ws["I5"].alignment = Alignment(horizontal="center")

#cabeçalho e titulo
ws["D1"] = 'ESCALA PORTARIA {}'.format(year)
ws["D1"].alignment = Alignment(horizontal="center", vertical="center")
ws["A5"] = 'DATA'
ws["C5"] = 'FUNCIONARIO'
ws["E5"] = 'FUNCIONARIO'
ws["G5"] = 'FUNCIONARIO'
ws["I5"] = 'OBSERVAÇÃO'

for row in ws.iter_rows(min_row=5, min_col=1, max_row=5, max_col=11):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(size=13, bold=True)

#preenchimento de datas
data_inicio = datetime.strptime('01/01/{}'.format(year), "%d/%m/%Y").date()

for i in range(6,372):    
    ws.merge_cells("A{}:B{}".format(i,i))
    ws.merge_cells("C{}:D{}".format(i,i))
    ws.merge_cells("E{}:F{}".format(i,i))
    ws.merge_cells("G{}:H{}".format(i,i))
    ws.merge_cells("I{}:K{}".format(i,i))

for row in ws.iter_rows(min_row=6, min_col=1, max_row=371, max_col=11):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")

for i in range(6,372):    
    ws["A{}".format(i)] = data_inicio.strftime("%d/%m/%Y")
    data_inicio += timedelta(days=1)

#colocando bordas
for row in ws.iter_rows(min_row=1, min_col=1, max_row=371, max_col=11):
    for cell in row:
        cell.border = thin_border

#editando fontes e cores 
c = ws["D1"]
c.font = Font(size=20, bold=True, color="FFFFFF")
c.fill = fill = PatternFill("solid", fgColor="1E90FF")


for col in ws.iter_cols(min_row=1, min_col = 12, max_row=371, max_col = 142):
    ws.column_dimensions[col].hidden= True

#salvando arquivo
wb.save('Portaria_Escala_{}.xlsx'.format(year))
